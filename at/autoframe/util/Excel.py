# -*- coding:utf-8 -*-

from openpyxl import *
from openpyxl.styles import fonts
from FormatTime import data_time_chinese

class ParseExcel(object):
    def __init__(self,excel_file_path):
        self.excel_file_path=excel_file_path
        self.workbook=load_workbook(excel_file_path)
        self.font=fonts(color=None)
        self.colorDict={"red":'FFFF3030',"green":'FF008B00'}
        self.sheet=self.get_sheet_by_index(0)

    # 通过序号设置当前要操作的sheet，使用index来获取相应的sheet
    def set_sheet_by_index(self,sheet_index):
        self.sheet = self.get_sheet_by_index(sheet_index)

    # 通过名字设置操作的sheet
    def set_sheet_by_name(self,sheet_name):
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)

    # 获取当前操作的sheet和title名字
    def get_default_name(self):
        return self.sheet.title

    # 通过名字获取要操作的sheet
    def get_sheet_by_name(self,sheet_name):
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)
        return self.sheet

    # 通过序号获取要操作的sheet
    def get_sheet_by_index(self, sheet_index):
        sheet_name = self.workbook.get_sheet_names()[sheet_index]
        self.sheet = self.get_sheet_by_name(sheet_name)
        return self.sheet

    #获取sheet中的最大行号，从0开始
    def get_max_row_no(self):
        return self.sheet.max_row

    # 获取sheet中的最大列号，从0开始
    def get_max_col_no(self):
        return self.sheet.max_column

    #获取默认sheet中的最小行号
    def get_min_row_no(self):
        return self.sheet.min_row

    # 获取默认sheet中的最小列号
    def get_min_col_no(self):
        return self.sheet.min_column

    #获取正在操作的sheet中的所有行对象
    def get_all_rows(self):
        # rows = []
        # for row in self.sheet.iter_rows():
        #     rows.append(row)
        # return rows
        #也可用以上方法
        return list(self.sheet.iter_rows())

    #获取正在操作的sheet中的所有列对象
    def get_all_cols(self):
        # cols = []
        # for col in self.sheet.iter_cols():
        #     cols.append(col)
        # return cols
        #也可用以上方法
        return list(self.sheet.iter_cols())

    #获取某一个行对象，第一行从0开始
    def get_single_row(self,row_no):
        return self.get_all_rows()[row_no]

    # 获取某一个列对象，第一列从0开始
    def get_single_col(self, col_no):
        return self.get_all_cols()[col_no]

    #获取某一个单元格对象,行号和列号从1开始
    def get_cell(self,row_no,col_no):
        return self.sheet.cell(row = row_no,column = col_no)

    # 获取某一个单元格内容
    def get_cell_content(self, row_no, col_no):
        return self.sheet.cell(row = row_no,column = col_no).value

    # 给某一个单元格写入指定内容,行号和列号从1开始
    #调用此方法时，excel不要处于打开状态
    def write_cell_content(self, row_no, col_no,content,font=None):
        self.sheet.cell(row=row_no,column=col_no).value = content
        self.workbook.save(self.excel_file_path)

    #给某个单元格写入当前时间,行号和列号从1开始
    #调用此方法时，excel不要处于打开状态
    def write_cell_current_time(self,row_no,col_no):
        self.sheet.cell(row=row_no,column=col_no).value = data_time_chinese()
        self.workbook.save(self.excel_file_path)

    def save_excel_file(self):
        self.workbook.save(self.excel_file_path)
    #保存所有对单元格的修改

if __name__ == "__main__":
    #测试所有的方法
    pe = ParseExcel(r"../TestData/联系人.xlsx")
    pe.set_sheet_by_index(0)
    print pe.get_default_name()
    pe.set_sheet_by_name("2")
    print pe.get_default_name()
    print pe.get_sheet_by_name("2")
    print pe.get_sheet_by_index(0)

    print "max rows:",pe.get_max_row_no()
    print "min row",pe.get_min_row_no()
    print pe.get_all_rows() #获取所有行对象
    print pe.get_all_rows()[2] #获取某一行
    print pe.get_all_rows()[0][2] #获取某一个单元格
    print pe.get_all_rows()[2][1].value #获取某一单元格的值

    print "max cols:",pe.get_max_col_no()
    print "min col",pe.get_min_col_no()
    print pe.get_all_cols() #获取所有行对象
    print pe.get_all_cols()[2] #获取某一行
    print pe.get_all_cols()[0][2] #获取某一个单元格
    print pe.get_all_cols()[2][1].value #获取某一单元格的值

    print len(pe.get_all_rows())
    for cell in pe.get_all_rows()[1]:
        print cell.value

    print len(pe.get_all_cols())
    for cell in pe.get_all_cols()[2]:
        print cell.value

    print "================================"

    for cell in pe.get_single_col(0):
        print cell.value

    for cell in pe.get_single_row(0):
        print cell.value

    print "--------------------"
    print pe.get_cell(1,2)
    print pe.get_cell_content(1,1)

    pe.write_cell_content(5,6,"hello")
    print pe.get_cell_content(5,6)
    pe.write_cell_current_time(7,7)
    print pe.get_cell_content(7,7)